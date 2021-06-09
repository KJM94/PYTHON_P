'''
[모듈과 패키지]
    모듈(Module)
        -변수,함수,클래스 등을 모아 놓은 소스파일
        -간단한 기능을 담을 때 사용
        파이썬에서는 파일 하나하나가 모듈
            -다른 파이썬 프로그램에서 불러와 사용할 수 있도록 만들어진 소스파일

    패키지(Package)
        -여러 모듈을 묶은 것
        -코드가 많고 복잡할 때 사용
        -관련 모듈(소스파일)들 끼리 한 폴더에 넣어놓은 형태(폴더 == 패키지)

        - 폴더 (패키지) 안에 __init__.py 파일이 있어야 패키지가 됐었다
        파이썬 3.3 버전 부터는 없어도 패키지로 인식
        *(하위 버전 호환을 위해 가급적 작성)

        
            
'''

'''
#모듈 불러오기(1) - 모듈명 그대로 사용
import koreais_module

print(koreais_module.my_str)
print(koreais_module.add(1,2))
print(koreais_module.sub(1,2))
print(koreais_module.mul(1,2))
print(koreais_module.div(1,2))

baduk=koreais_module.Dog("바둑")
baduk.cry()

#모듈 불러오기 (2) - as 별칭 주기
import koreais_module as km

print(km.my_str)

#모듈 불러오기 (3) - 모듈에서 일부만 사용
from koreais_module import my_str,add
#add() 이렇게 쓰지 않고 add로 사용(이름만 사용)

print(my_str)
print(add(1,2))
#print(sub(2,1))

#모듈 불러오기 (4) - 모듈에서 전체 불러오기
from koreais_module import *
print(sub(2,1))

#모듈을 import 할 때 모듈명을 생략하는 3,4번 방식은 좋지 않다.
#왜 ? 모듈은 여러개 import 하다보면 각 모듈에서 이름이 중복 발생 가능성이 존재
#가장 좋은 방법은 2번 별칭 주기
'''

#sys.path 안에 모듈이 위치해야 한다.
#sys.path.append("c:\\...\\") 형태로 경로 추가 가능

import sys
print(sys.path)

#import subprocess
#print(subprocess.check_butput("dir",shell=True))

#help("modules")

#미리 정의된 함수 집합 = 라이브러리
import math
num = math.sqrt(5)
print(num)

num = math.factorial(5)
print(num)


#해당 라이브러리에 있는 함수를 사용하기 위해서는 해당 라이브러리명.함수명
#하나의 라이브러리에는 여러 함수 존재
#라이브러리가 없으면 실행 x --> 설치 하면 됨

#pip install 라이브러리
#pip search 찾을 키워드
#pip uninstall 삭제할 패키지

#<파이썬 설치 경로>\Lib\site-pakage 경로 설치된

'''
import request
url = "http:\\www.naver.com"
res = request.get(url)
print(res.content)
'''

#openpyxl 설치 -> 액셀 파일을 다루는 라이브러리
#pip install openpyxl
'''
anaconda 패키지를 설치하면 openpyxl 라이브러리 설치 x
anaconda 패키지:
파이썬 컴파일러 + 주요라이브러리 + 주요툴 모음
'''

import openpyxl
'''
wb = openpyxl.load_workbook(".\data\train.xlsx")
sheet = wb['sheet1']
print(sheet.cell(row=1,column=1).value)
print("%s"%sheet.max=column)
'''
'''
wb = openpyxl.Workbook() #workbook함수를 통해 객체 생성
#생성시 기본적으로 sheet1이 생성

sheet2 = wb.create_sheet('sheet2') #마지막에 추가
sheet3 = wb.create_sheet('sheet3',1) #sheet1 자리에 삽입하여 추가

sheet2.title = '업무자동화' #sheet 이름 변경

sheet2['A1'] = 'hello'

print(sheet2.cell(row=1,column=1).value)


wb.save(".\new_test_file.xlsx") #파일 저장




wb.close() #파일 닫기
'''

'''
여러 cell 접근 하기
(1) 특정 범위
    -cell_range = sheet['A1':'C2']
    특정 row
    row10 = sheet[10]

    특정 row 범위
    row_range = sheet[5:10]

    특정 column
    colc = sheet['C']

    특정 column 범위
    col_range = sheet['C:D']
'''
import requests
url = "http://www.naver.com"
res = requests.get(url)
print(res.content)





























